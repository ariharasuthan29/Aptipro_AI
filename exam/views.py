import io
import json
import random
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Avg, Max, Count, Sum, Q
import openpyxl

from .models import (
    UserProfile, Category, Question, ExamSession, 
    ExamQuestionResponse, ViolationLog, DailyActivity
)
from .proctoring import analyze_webcam_frame

# --- Helper Functions ---
def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

def track_daily_login(user):
    today = timezone.now().date()
    profile, _ = UserProfile.objects.get_or_create(user=user)
    
    activity, created = DailyActivity.objects.get_or_create(user=user, date=today)
    
    last_date = profile.last_login_date
    if hasattr(last_date, 'date'):
        last_date = last_date.date()

    if last_date != today:
        if last_date == today - timedelta(days=1):
            profile.daily_streak += 1
        elif last_date < today - timedelta(days=1):
            profile.daily_streak = 1
        profile.last_login_date = today
        profile.save()

# --- Authentication Views ---
def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        phone = request.POST.get('phone', '').strip()
        target_exam = request.POST.get('target_exam', 'Placement Examination').strip()
        
        if not username or not password:
            messages.error(request, "Username and password are required.")
            return render(request, 'register.html')
            
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'register.html')
            
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username is already taken.")
            return render(request, 'register.html')
            
        user = User.objects.create_user(username=username, email=email, password=password)
        UserProfile.objects.create(user=user, phone=phone, target_exam=target_exam)
        
        login(request, user)
        track_daily_login(user)
        messages.success(request, f"Welcome to OnlineExamSystem, {user.username}!")
        return redirect('dashboard')
        
    return render(request, 'register.html')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            track_daily_login(user)
            messages.success(request, f"Welcome back, {user.username}!")
            if user.is_staff or user.is_superuser:
                return redirect('admin_dashboard')
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
            
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


# --- Student Dashboard & Analytics Views ---
@login_required
def dashboard_view(request):
    track_daily_login(request.user)
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    
    # Completed exam sessions
    sessions = ExamSession.objects.filter(user=user, status__in=[ExamSession.COMPLETED, ExamSession.AUTO_SUBMITTED]).order_by('-start_time')
    
    total_tests = sessions.count()
    avg_score = sessions.aggregate(Avg('percentage'))['percentage__avg'] or 0.0
    highest_score = sessions.aggregate(Max('percentage'))['percentage__max'] or 0.0
    
    # Recent 5 attempts
    recent_exams = sessions[:5]
    
    # Score History Graph Data (Last 10 exams)
    chart_labels = []
    chart_scores = []
    for s in reversed(sessions[:10]):
        chart_labels.append(s.start_time.strftime("%b %d, %H:%M"))
        chart_scores.append(round(s.percentage, 1))
        
    # Category Breakdown Performance
    categories = Category.objects.all()
    cat_performance = []
    for cat in categories:
        resp = ExamQuestionResponse.objects.filter(exam_session__user=user, question__category=cat)
        total_q = resp.count()
        correct_q = resp.filter(is_correct=True).count()
        acc = round((correct_q / total_q * 100), 1) if total_q > 0 else 0
        cat_performance.append({
            'name': cat.name,
            'icon': cat.icon,
            'total': total_q,
            'correct': correct_q,
            'accuracy': acc
        })
        
    # Update profile stats
    profile.total_exams = total_tests
    profile.average_score = round(avg_score, 1)
    profile.highest_score = round(highest_score, 1)
    profile.save()

    context = {
        'profile': profile,
        'total_tests': total_tests,
        'avg_score': round(avg_score, 1),
        'highest_score': round(highest_score, 1),
        'recent_exams': recent_exams,
        'chart_labels': json.dumps(chart_labels),
        'chart_scores': json.dumps(chart_scores),
        'cat_performance': cat_performance,
        'categories': categories,
    }
    return render(request, 'dashboard.html', context)


# --- Daily Practice Module ---
@login_required
def practice_view(request):
    category_id = request.GET.get('category')
    difficulty = request.GET.get('difficulty')
    
    questions = Question.objects.all()
    selected_category = None
    
    if category_id:
        selected_category = get_object_or_404(Category, id=category_id)
        questions = questions.filter(category=selected_category)
    if difficulty:
        questions = questions.filter(difficulty=difficulty)
        
    questions = list(questions)
    random.shuffle(questions)
    questions = questions[:10]  # 10 Questions for daily practice drill
    
    context = {
        'questions': questions,
        'categories': Category.objects.all(),
        'selected_category': selected_category,
        'selected_difficulty': difficulty,
    }
    return render(request, 'practice.html', context)


# --- Randomized Examination Module & Anti-Cheating System ---
@login_required
def exam_start_view(request):
    # Check if active uncompleted session exists
    active_session = ExamSession.objects.filter(user=request.user, status=ExamSession.IN_PROGRESS).first()
    if active_session:
        return redirect('exam_room', session_key=active_session.session_key)
        
    categories_count = Category.objects.annotate(q_count=Count('questions'))
    total_q_in_db = Question.objects.count()
    
    context = {
        'total_questions_db': total_q_in_db,
        'categories_count': categories_count,
    }
    return render(request, 'exam_start.html', context)

@login_required
def start_exam_session_view(request):
    if request.method != 'POST':
        return redirect('exam_start')
        
    # Check if active uncompleted session already exists
    active_session = ExamSession.objects.filter(user=request.user, status=ExamSession.IN_PROGRESS).first()
    if active_session:
        return redirect('exam_room', session_key=active_session.session_key)
        
    # Random selection: 20 Easy, 15 Medium, 5 Hard (Total 40 MCQs)
    easy_pool = list(Question.objects.filter(difficulty=Question.EASY))
    med_pool = list(Question.objects.filter(difficulty=Question.MEDIUM))
    hard_pool = list(Question.objects.filter(difficulty=Question.HARD))
    
    selected_easy = random.sample(easy_pool, min(20, len(easy_pool)))
    selected_med = random.sample(med_pool, min(15, len(med_pool)))
    selected_hard = random.sample(hard_pool, min(5, len(hard_pool)))
    
    selected_questions = selected_easy + selected_med + selected_hard
    
    # Fallback if DB has fewer questions
    if len(selected_questions) < 40:
        all_pool = list(Question.objects.exclude(id__in=[q.id for q in selected_questions]))
        needed = 40 - len(selected_questions)
        selected_questions += random.sample(all_pool, min(needed, len(all_pool)))
        
    random.shuffle(selected_questions)
    
    # Create ExamSession
    session = ExamSession.objects.create(
        user=request.user,
        mode=ExamSession.FULL_EXAM,
        status=ExamSession.IN_PROGRESS,
        duration_seconds=3600,
        remaining_seconds=3600,
        total_questions=len(selected_questions),
        start_time=timezone.now()
    )
    
    # Populate ExamQuestionResponse with randomized option shuffling
    for q in selected_questions:
        options = [('A', q.option_a), ('B', q.option_b), ('C', q.option_c), ('D', q.option_d)]
        random.shuffle(options)
        
        # Mapping shuffled position (A,B,C,D) to original option letter (A,B,C,D) and text
        shuffled_mapping = {
            'A': {'orig_key': options[0][0], 'text': options[0][1]},
            'B': {'orig_key': options[1][0], 'text': options[1][1]},
            'C': {'orig_key': options[2][0], 'text': options[2][1]},
            'D': {'orig_key': options[3][0], 'text': options[3][1]},
        }
        
        ExamQuestionResponse.objects.create(
            exam_session=session,
            question=q,
            shuffled_options=shuffled_mapping
        )
        
    return redirect('exam_room', session_key=session.session_key)

@login_required
def exam_room_view(request, session_key):
    session = get_object_or_404(ExamSession, session_key=session_key, user=request.user)
    
    if session.status != ExamSession.IN_PROGRESS:
        messages.info(request, "This exam session has already been submitted.")
        return redirect('exam_result', session_key=session.session_key)
        
    # Calculate synced remaining time
    elapsed = (timezone.now() - session.start_time).total_seconds()
    remaining = max(0, session.duration_seconds - int(elapsed))
    
    if remaining <= 0:
        return finish_exam_view(request, session.session_key, auto_submitted=True, reason="Time Expired")
        
    session.remaining_seconds = remaining
    session.save()
    
    responses = session.responses.select_related('question', 'question__category').all()
    
    context = {
        'session': session,
        'responses': responses,
        'remaining_seconds': remaining,
    }
    return render(request, 'exam_room.html', context)


# --- Ajax APIs for Proctoring & Answers ---
@login_required
def submit_exam_response_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            response_id = data.get('response_id')
            selected_option = data.get('selected_option')  # 'A', 'B', 'C', 'D' or null
            
            resp = ExamQuestionResponse.objects.get(id=response_id, exam_session__user=request.user)
            if resp.exam_session.status != ExamSession.IN_PROGRESS:
                return JsonResponse({'status': 'error', 'message': 'Exam already finished'}, status=400)
                
            resp.selected_option = selected_option
            
            # Check correctness against original correct_option
            if selected_option and resp.shuffled_options and selected_option in resp.shuffled_options:
                orig_key = resp.shuffled_options[selected_option]['orig_key']
                resp.is_correct = (orig_key == resp.question.correct_option)
            else:
                resp.is_correct = False
                
            resp.save()
            return JsonResponse({'status': 'success', 'response_id': resp.id, 'selected': selected_option})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)

@login_required
def analyze_frame_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            session_key = data.get('session_key')
            frame_base64 = data.get('frame')
            
            session = get_object_or_404(ExamSession, session_key=session_key, user=request.user)
            if session.status != ExamSession.IN_PROGRESS:
                return JsonResponse({'status': 'terminated', 'message': 'Exam session ended'})
                
            analysis = analyze_webcam_frame(frame_base64)
            
            if analysis['status'] == 'VIOLATION':
                # Increment violation counter
                session.violation_count += 1
                warning_level = min(3, session.violation_count)
                session.save()
                
                v_log = ViolationLog.objects.create(
                    exam_session=session,
                    user=request.user,
                    violation_type=analysis['violation_type'],
                    warning_level=warning_level,
                    details=analysis['details']
                )
                if analysis['image_file']:
                    v_log.snapshot.save(analysis['image_file'].name, analysis['image_file'], save=True)
                    
                auto_terminate = (session.violation_count >= 3)
                
                return JsonResponse({
                    'status': 'violation',
                    'violation_type': analysis['violation_type'],
                    'details': analysis['details'],
                    'warning_level': warning_level,
                    'violation_count': session.violation_count,
                    'auto_terminate': auto_terminate
                })
                
            return JsonResponse({'status': 'ok', 'faces_count': analysis['faces_count']})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@login_required
def log_violation_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            session_key = data.get('session_key')
            violation_type = data.get('violation_type')
            details = data.get('details', '')
            
            session = get_object_or_404(ExamSession, session_key=session_key, user=request.user)
            if session.status != ExamSession.IN_PROGRESS:
                return JsonResponse({'status': 'terminated'})
                
            session.violation_count += 1
            warning_level = min(3, session.violation_count)
            session.save()
            
            ViolationLog.objects.create(
                exam_session=session,
                user=request.user,
                violation_type=violation_type,
                warning_level=warning_level,
                details=details
            )
            
            auto_terminate = (session.violation_count >= 3)
            
            return JsonResponse({
                'status': 'success',
                'violation_count': session.violation_count,
                'warning_level': warning_level,
                'auto_terminate': auto_terminate
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)


# --- Exam Finish & Result Module ---
@login_required
def finish_exam_view(request, session_key, auto_submitted=False, reason=None):
    session = get_object_or_404(ExamSession, session_key=session_key, user=request.user)
    
    if session.status == ExamSession.IN_PROGRESS:
        responses = session.responses.all()
        correct_cnt = responses.filter(is_correct=True).count()
        answered_cnt = responses.exclude(selected_option__isnull=True).exclude(selected_option='').count()
        incorrect_cnt = answered_cnt - correct_cnt
        unanswered_cnt = session.total_questions - answered_cnt
        
        score = float(correct_cnt)
        percentage = round((correct_cnt / session.total_questions) * 100, 2) if session.total_questions > 0 else 0.0
        
        time_taken = max(0, session.duration_seconds - session.remaining_seconds)
        
        session.score = score
        session.percentage = percentage
        session.correct_answers_count = correct_cnt
        session.incorrect_answers_count = incorrect_cnt
        session.unanswered_count = unanswered_cnt
        session.time_taken_seconds = time_taken
        session.end_time = timezone.now()
        
        if auto_submitted or session.violation_count >= 3:
            session.status = ExamSession.AUTO_SUBMITTED
        else:
            session.status = ExamSession.COMPLETED
            
        session.save()
        
        # Award profile points
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        profile.total_points += int(score * 10)
        profile.save()
        
    return redirect('exam_result', session_key=session.session_key)

@login_required
def result_view(request, session_key):
    session = get_object_or_404(ExamSession, session_key=session_key, user=request.user)
    responses = session.responses.select_related('question', 'question__category').all()
    violations = session.violations.all().order_by('timestamp')
    
    # Difficulty accuracy analysis
    diff_stats = {}
    for level in [Question.EASY, Question.MEDIUM, Question.HARD]:
        level_resp = responses.filter(question__difficulty=level)
        t_cnt = level_resp.count()
        c_cnt = level_resp.filter(is_correct=True).count()
        diff_stats[level] = {
            'total': t_cnt,
            'correct': c_cnt,
            'acc': round((c_cnt / t_cnt * 100), 1) if t_cnt > 0 else 0
        }
        
    # Category accuracy analysis
    cat_stats = {}
    for resp in responses:
        c_name = resp.question.category.name
        if c_name not in cat_stats:
            cat_stats[c_name] = {'total': 0, 'correct': 0}
        cat_stats[c_name]['total'] += 1
        if resp.is_correct:
            cat_stats[c_name]['correct'] += 1
            
    for c_name in cat_stats:
        t = cat_stats[c_name]['total']
        c = cat_stats[c_name]['correct']
        cat_stats[c_name]['acc'] = round((c / t * 100), 1) if t > 0 else 0

    context = {
        'session': session,
        'responses': responses,
        'violations': violations,
        'diff_stats': diff_stats,
        'cat_stats': cat_stats,
        'time_taken_min': session.time_taken_seconds // 60,
        'time_taken_sec': session.time_taken_seconds % 60,
    }
    return render(request, 'result.html', context)


# --- Leaderboard Module ---
@login_required
def leaderboard_view(request):
    timeframe = request.GET.get('timeframe', 'overall') # overall, monthly, weekly
    
    users = User.objects.all().select_related('profile')
    leaderboard_data = []
    
    now = timezone.now()
    if timeframe == 'weekly':
        start_date = now - timedelta(days=7)
    elif timeframe == 'monthly':
        start_date = now - timedelta(days=30)
    else:
        start_date = None
        
    for u in users:
        qs = ExamSession.objects.filter(user=u, status__in=[ExamSession.COMPLETED, ExamSession.AUTO_SUBMITTED])
        if start_date:
            qs = qs.filter(start_time__gte=start_date)
            
        tests_count = qs.count()
        if tests_count == 0 and timeframe != 'overall':
            continue
            
        avg_pct = qs.aggregate(Avg('percentage'))['percentage__avg'] or 0.0
        max_pct = qs.aggregate(Max('percentage'))['percentage__max'] or 0.0
        points = getattr(u, 'profile', None).total_points if hasattr(u, 'profile') else 0
        streak = getattr(u, 'profile', None).daily_streak if hasattr(u, 'profile') else 0
        
        leaderboard_data.append({
            'username': u.username,
            'tests_count': tests_count,
            'avg_pct': round(avg_pct, 1),
            'max_pct': round(max_pct, 1),
            'points': points,
            'streak': streak
        })
        
    # Sort by points descending then avg_pct descending
    leaderboard_data.sort(key=lambda x: (x['points'], x['avg_pct']), reverse=True)
    
    # Assign ranks
    for rank, item in enumerate(leaderboard_data, start=1):
        item['rank'] = rank
        
    context = {
        'leaderboard': leaderboard_data,
        'timeframe': timeframe,
    }
    return render(request, 'leaderboard.html', context)


# --- Admin Panel & Excel Question Importer ---
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    total_students = User.objects.filter(is_staff=False).count()
    total_questions = Question.objects.count()
    total_exams = ExamSession.objects.count()
    total_violations = ViolationLog.objects.count()
    
    recent_violations = ViolationLog.objects.select_related('user', 'exam_session').order_by('-timestamp')[:10]
    recent_exams = ExamSession.objects.select_related('user').order_by('-start_time')[:10]
    
    context = {
        'total_students': total_students,
        'total_questions': total_questions,
        'total_exams': total_exams,
        'total_violations': total_violations,
        'recent_violations': recent_violations,
        'recent_exams': recent_exams,
    }
    return render(request, 'admin_dashboard.html', context)

@user_passes_test(is_admin)
def admin_questions_view(request):
    category_id = request.GET.get('category')
    difficulty = request.GET.get('difficulty')
    search_q = request.GET.get('search', '').strip()
    
    questions = Question.objects.select_related('category').all().order_by('-created_at')
    
    if category_id:
        questions = questions.filter(category_id=category_id)
    if difficulty:
        questions = questions.filter(difficulty=difficulty)
    if search_q:
        questions = questions.filter(Q(question_text__icontains=search_q) | Q(category__name__icontains=search_q))
        
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            cat_id = request.POST.get('category')
            diff = request.POST.get('difficulty')
            q_text = request.POST.get('question_text')
            op_a = request.POST.get('option_a')
            op_b = request.POST.get('option_b')
            op_c = request.POST.get('option_c')
            op_d = request.POST.get('option_d')
            correct = request.POST.get('correct_option')
            exp = request.POST.get('explanation')
            
            cat = get_object_or_404(Category, id=cat_id)
            Question.objects.create(
                category=cat, difficulty=diff, question_text=q_text,
                option_a=op_a, option_b=op_b, option_c=op_c, option_d=op_d,
                correct_option=correct, explanation=exp
            )
            messages.success(request, "New question added successfully.")
            return redirect('admin_questions')
            
        elif action == 'delete':
            q_id = request.POST.get('question_id')
            Question.objects.filter(id=q_id).delete()
            messages.success(request, "Question deleted.")
            return redirect('admin_questions')

    context = {
        'questions': questions,
        'categories': Category.objects.all(),
        'selected_category': category_id,
        'selected_difficulty': difficulty,
        'search_q': search_q,
    }
    return render(request, 'admin_questions.html', context)

@user_passes_test(is_admin)
def admin_violations_view(request):
    violations = ViolationLog.objects.select_related('user', 'exam_session').order_by('-timestamp')
    context = {
        'violations': violations
    }
    return render(request, 'admin_violations.html', context)

@user_passes_test(is_admin)
def export_sample_excel_template_view(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Question Bank Template"
    
    headers = ["Category", "Difficulty", "Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Option", "Explanation"]
    ws.append(headers)
    
    sample_rows = [
        ["Quantitative Aptitude", "EASY", "What is 12 + 15?", "25", "27", "30", "22", "B", "12 + 15 = 27."],
        ["Logical Reasoning", "MEDIUM", "Find next in series: 2, 4, 8, 16, ?", "24", "32", "64", "20", "B", "Multiply by 2 each step."],
        ["Programming Basics", "HARD", "What is the worst-case time complexity of QuickSort?", "O(N log N)", "O(N)", "O(N^2)", "O(1)", "C", "Degrades to O(N^2) on bad pivot choices."]
    ]
    for row in sample_rows:
        ws.append(row)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=Aptitude_Questions_Sample_Template.xlsx'
    wb.save(response)
    return response

@user_passes_test(is_admin)
def import_excel_questions_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        
        if not (file.name.endswith('.xlsx') or file.name.endswith('.csv')):
            messages.error(request, "Please upload a valid .xlsx or .csv file.")
            return redirect('admin_questions')
            
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            
            imported_count = 0
            skipped_count = 0
            
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) <= 1:
                messages.warning(request, "Excel file appears to be empty or has only headers.")
                return redirect('admin_questions')
                
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue
                    
                cat_name = str(row[0]).strip() if len(row) > 0 and row[0] else 'General Aptitude'
                diff_val = str(row[1]).strip().upper() if len(row) > 1 and row[1] else 'EASY'
                q_text = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                op_a = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                op_b = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                op_c = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                op_d = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                correct = str(row[7]).strip().upper() if len(row) > 7 and row[7] else 'A'
                explanation = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                
                if not q_text or not op_a or not op_b:
                    skipped_count += 1
                    continue
                    
                if diff_val not in ['EASY', 'MEDIUM', 'HARD']:
                    diff_val = 'EASY'
                if correct not in ['A', 'B', 'C', 'D']:
                    correct = 'A'
                    
                cat_obj, _ = Category.objects.get_or_create(
                    name=cat_name,
                    defaults={'slug': f"{slugify(cat_name)}-{row_idx}", 'icon': 'bi-journal-check'}
                )
                
                Question.objects.create(
                    category=cat_obj,
                    difficulty=diff_val,
                    question_text=q_text,
                    option_a=op_a,
                    option_b=op_b,
                    option_c=op_c,
                    option_d=op_d,
                    correct_option=correct,
                    explanation=explanation
                )
                imported_count += 1
                
            messages.success(request, f"Successfully imported {imported_count} questions from Excel! (Skipped: {skipped_count})")
        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")
            
        return redirect('admin_questions')
        
    return redirect('admin_questions')


@login_required
def submit_practice_view(request):
    if request.method != 'POST':
        return redirect('practice')
        
    q_ids = request.POST.getlist('question_ids')
    responses = []
    correct_count = 0
    total_count = len(q_ids)
    
    for q_id in q_ids:
        q = get_object_or_404(Question, id=q_id)
        selected = request.POST.get(f'answer_{q_id}', '')
        is_correct = (selected == q.correct_option)
        if is_correct:
            correct_count += 1
            
        responses.append({
            'question': q,
            'selected_option': selected,
            'is_correct': is_correct,
        })
        
    # Calculate percentage
    pct = round((correct_count / total_count * 100), 1) if total_count > 0 else 0.0
    
    # Award points: +2 points per correct answer!
    profile = request.user.profile
    points_gained = correct_count * 2
    profile.points += points_gained
    profile.save()
    
    context = {
        'responses': responses,
        'correct_count': correct_count,
        'total_count': total_count,
        'percentage': pct,
        'points_gained': points_gained,
        'total_points': profile.points,
    }
    return render(request, 'practice_result.html', context)






